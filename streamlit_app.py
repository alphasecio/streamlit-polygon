import streamlit as st
import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression, Ridge, Lasso
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import TimeSeriesSplit, cross_val_score
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from scipy import stats
from statsmodels.tsa.seasonal import seasonal_decompose
from statsmodels.tsa.stattools import adfuller
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from dataclasses import dataclass
from typing import Optional, Dict, Any, List, Union
import io
import zipfile
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

# Suppress warnings
warnings.filterwarnings('ignore')

# Set page layout
st.set_page_config(layout="wide")

# Custom CSS styling
st.markdown("""
    <style>
    .stmetric .stmetriclabel {
        font-size: 14px !important;
        color: #555;
    }
    .stmetric .stmetricvalue {
        font-size: 24px !important;
        font-weight: bold;
        color: #0f3460;
    }
    .stmetric .stmetricdelta {
        font-size: 12px !important;
    }

    .required-field label::after {
        content: ' *';
        color: red;
        font-weight: bold;
    }

    .optional-field {
        opacity: 0.8;
    }
    .optional-field::after {
        content: ' (optional)';
        font-size: 0.8em;
        color: #666;
    }

    .success-message {
        padding: 1rem;
        background-color: #d4edda;
        border-color: #c3e6cb;
        color: #155724;
        border-radius: 0.25rem;
        margin: 1rem 0;
    }

    .warning-message {
        padding: 1rem;
        background-color: #fff3cd;
        border-color: #ffeeba;
        color: #856404;
        border-radius: 0.25rem;
        margin: 1rem 0;
    }

    .error-message {
        padding: 1rem;
        background-color: #f8d7da;
        border-color: #f5c6cb;
        color: #721c24;
        border-radius: 0.25rem;
        margin: 1rem 0;
    }

    .section-header {
        padding: 1rem 0;
        margin: 1rem 0;
        border-bottom: 2px solid #eee;
        color: #0f3460;
    }
    </style>
""", unsafe_allow_html=True)

# Constants
file_size_limit = 100 * 1024 * 1024  # 100mb in bytes
required_columns = ['month', 'leads', 'appointments', 'closings', 'cost']
numeric_columns = ['leads', 'appointments', 'closings', 'cost']

# State management classes
@dataclass
class AppState:
    initialized: bool = False
    data_loaded: bool = False
    current_tab: str = "input & upload"
    last_update: datetime = datetime.now()

    def to_dict(self) -> Dict[str, Any]:
        return {
            'initialized': self.initialized,
            'data_loaded': self.data_loaded,
            'current_tab': self.current_tab,
            'last_update': self.last_update.isoformat()
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'AppState':
        try:
            data['last_update'] = datetime.fromisoformat(data['last_update'])
            return cls(**data)
        except Exception as e:
            st.error(f"Error restoring app state: {str(e)}")
            return cls()

@dataclass
class DataState:
    historical_data: Optional[pd.DataFrame] = None
    filtered_data: Optional[pd.DataFrame] = None
    date_range: Optional[tuple] = None
    metrics: Optional[Dict] = None
    forecasts: Optional[Dict] = None

    def has_data(self) -> bool:
        return isinstance(self.historical_data, pd.DataFrame) and not self.historical_data.empty

# Utility functions for displaying messages
def show_success(message: str):
    st.markdown(f'<div class="success-message">{message}</div>', unsafe_allow_html=True)

def show_warning(message: str):
    st.markdown(f'<div class="warning-message">{message}</div>', unsafe_allow_html=True)

def show_error(message: str):
    st.markdown(f'<div class="error-message">{message}</div>', unsafe_allow_html=True)

# Enhanced forecasting class with multiple models and validation
class EnhancedForecaster:
    def __init__(self):
        self.models = {
            'linear regression': LinearRegression(),
            'ridge regression': Ridge(alpha=1.0),
            'lasso regression': Lasso(alpha=1.0),
            'random forest': RandomForestRegressor(n_estimators=100, random_state=42)
        }
        self.selected_model = None
        self.x_train = None
        self.y_train = None
        self.feature_importance = None

    def prepare_features(self, df: pd.DataFrame) -> pd.DataFrame:
        try:
            df = df.copy()
            df['month_num'] = df.index.month
            df['quarter'] = df.index.quarter
            df['year'] = df.index.year
            for lag in [1, 2, 3]:
                df[f'leads_lag_{lag}'] = df['leads'].shift(lag)
                df[f'appointments_lag_{lag}'] = df['appointments'].shift(lag)
            for window in [3, 6, 12]:
                df[f'leads_rolling_{window}'] = df['leads'].rolling(window=window).mean()
                df[f'appointments_rolling_{window}'] = df['appointments'].rolling(window=window).mean()
            return df.dropna()
        except Exception as e:
            st.error(f"Error preparing features: {str(e)}")
            return pd.DataFrame()

    def train(self, df: pd.DataFrame, target: str = 'closings', model_type: str = 'linear regression') -> None:
        try:
            prepared_df = self.prepare_features(df)
            if prepared_df.empty:
                raise ValueError("No data available after feature preparation")

            feature_cols = [col for col in prepared_df.columns if col not in ['closings', 'month', 'cost']]
            self.x_train = prepared_df[feature_cols]
            self.y_train = prepared_df[target]
            self.selected_model = self.models[model_type]
            self.selected_model.fit(self.x_train, self.y_train)

            if model_type == 'random forest':
                self.feature_importance = pd.DataFrame({
                    'feature': feature_cols,
                    'importance': self.selected_model.feature_importances_
                }).sort_values('importance', ascending=False)
        except Exception as e:
            st.error(f"Error training model: {str(e)}")

    def cross_validate(self, n_splits: int = 5) -> np.ndarray:
        try:
            tscv = TimeSeriesSplit(n_splits=n_splits)
            return cross_val_score(self.selected_model, self.x_train, self.y_train, cv=tscv, scoring='r2')
        except Exception as e:
            st.error(f"Error in cross-validation: {str(e)}")
            return np.array([])

    def predict_with_intervals(self, x: pd.DataFrame, confidence: float = 0.95) -> tuple:
        try:
            predictions = self.selected_model.predict(x)
            y_pred_train = self.selected_model.predict(self.x_train)
            residuals = self.y_train - y_pred_train
            std_resid = np.std(residuals)

            z_score = stats.norm.ppf((1 + confidence) / 2)
            margin = z_score * std_resid

            lower_bound = predictions - margin
            upper_bound = predictions + margin

            return predictions, lower_bound, upper_bound
        except Exception as e:
            st.error(f"Error making predictions: {str(e)}")
            return np.array([]), np.array([]), np.array([])

class EnhancedExporter:
    def __init__(self, data_state: DataState, app_state: AppState):
        self.data_state = data_state
        self.app_state = app_state
        self.timestamp = datetime.now().strftime("%y%m%d_%H%M%S")

    def _calculate_summary_metrics(self) -> dict[str, Union[int, float, str]]:
        try:
            df = self.data_state.historical_data
            if df is None or df.empty:
                return {}
            return {
                "total_leads": int(df['leads'].sum()),
                "total_appointments": int(df['appointments'].sum()),
                "total_closings": int(df['closings'].sum()),
                "average_conversion_rate": float(df['closings'].sum() / df['leads'].sum() * 100 if df['leads'].sum() > 0 else 0),
                "best_performing_month": df.loc[df['closings'].idxmax(), 'month'].strftime("%y-%m") if not df['closings'].empty else 'n/a'
            }
        except Exception as e:
            st.error(f"Error calculating summary metrics: {str(e)}")
            return {}

    def generate_metadata(self) -> dict[str, Any]:
        try:
            if not isinstance(self.data_state.historical_data, pd.DataFrame):
                return {}
            return {
                "export_info": {
                    "generated_at": datetime.now().isoformat(),
                    "app_version": "1.0.0",
                    "user_settings": self.app_state.to_dict()
                },
                "data_summary": {
                    "date_range": {
                        "start": self.data_state.historical_data['month'].min().isoformat(),
                        "end": self.data_state.historical_data['month'].max().isoformat()
                    },
                    "record_count": len(self.data_state.historical_data),
                    "metrics_summary": self._calculate_summary_metrics(),
                    "forecast_settings": self.data_state.forecasts
                }
            }
        except Exception as e:
            st.error(f"Error generating metadata: {str(e)}")
            return {}

    def create_excel_report(self, include_sections: List[str]) -> bytes:
        try:
            output = io.BytesIO()
            workbook = Workbook()
            header_style = Font(bold=True, color="ffffff")
            header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")

            if 'historical_data' in include_sections:
                self._add_historical_data_sheet(workbook, header_style, header_fill)
            if 'metrics' in include_sections:
                self._add_metrics_sheet(workbook, header_style, header_fill)
            if 'forecasts' in include_sections and self.data_state.forecasts:
                self._add_forecasts_sheet(workbook, header_style, header_fill)
            if 'metadata' in include_sections:
                self._add_metadata_sheet(workbook, header_style, header_fill)

            workbook.save(output)
            return output.getvalue()
        except Exception as e:
            st.error(f"Error creating excel report: {str(e)}")
            return b''

    def _add_historical_data_sheet(self, workbook: Workbook, header_style: Font, header_fill: PatternFill) -> None:
        try:
            ws = workbook.create_sheet("historical data")
            df = self.data_state.historical_data

            for col_num, column in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num, value=column)
                cell.font = header_style
                cell.fill = header_fill

            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)

            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column[0].column_letter].width = max_length + 2
        except Exception as e:
            st.error(f"Error adding historical data sheet: {str(e)}")

    def create_zip_export(self, include_sections: List[str]) -> bytes:
        try:
            zip_buffer = io.BytesIO()

            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                excel_data = self.create_excel_report(include_sections)
                if excel_data:
                    zip_file.writestr(f'sales_forecast_{self.timestamp}.xlsx', excel_data)

                metadata = yaml.dump(self.generate_metadata(), default_flow_style=False)
                zip_file.writestr(f'metadata_{self.timestamp}.yaml', metadata)

                readme_content = self._generate_readme()
                zip_file.writestr('readme.txt', readme_content)

            return zip_buffer.getvalue()
        except Exception as e:
            st.error(f"Error creating zip export: {str(e)}")
            return b''

    def _generate_readme(self) -> str:
        return f"""Sales Forecast Export
Generated: {datetime.now().strftime("%y-%m-%d %H:%M:%S")}

This export package contains the following files:
1. Excel report (sales_forecast_{self.timestamp}.xlsx)
   - Historical data
   - Metrics analysis
   - Forecasts
   - Metadata

2. Metadata (metadata_{self.timestamp}.yaml)
   - Detailed export metadata in YAML format

For questions or support, please contact support@example.com
"""

# Cached analytics functions
@st.cache_data(ttl=3600)
def analyze_seasonality(df: pd.DataFrame, metric: str) -> tuple:
    try:
        decomposition = seasonal_decompose(df[metric], period=12, model='additive')

        fig = go.Figure()

        fig.add_trace(go.Scatter(x=df.index, y=decomposition.observed, name='observed', line=dict(color='blue')))
        fig.add_trace(go.Scatter(x=df.index, y=decomposition.trend, name='trend', line=dict(color='red')))
        fig.add_trace(go.Scatter(x=df.index, y=decomposition.seasonal, name='seasonal', line=dict(color='green')))
        fig.add_trace(go.Scatter(x=df.index, y=decomposition.resid, name='residual', line=dict(color='gray')))

        fig.update_layout(title=f'Decomposition of {metric}', height=800)
        return fig, decomposition
    except Exception as e:
        st.error(f"Error in seasonality analysis: {str(e)}")
        return None, None

@st.cache_data(ttl=3600)
def check_stationarity(series: pd.Series) -> dict[str, Union[float, bool]]:
    try:
        result = adfuller(series.dropna())
        return {
            'test_statistic': result[0],
            'p_value': result[1],
            'is_stationary': result[1] < 0.05
        }
    except Exception as e:
        st.error(f"Error in stationarity test: {str(e)}")
        return {'test_statistic': 0.0, 'p_value': 1.0, 'is_stationary': False}

# State management functions
def initialize_session_state() -> None:
    try:
        if 'app_state' not in st.session_state:
            st.session_state.app_state = AppState()

        if 'data_state' not in st.session_state:
            st.session_state.data_state = DataState()

        if not st.session_state.app_state.initialized:
            st.session_state.input_values = {
                'leads': 100,
                'appointments': 50,
                'closings': 25,
                'cost': 0.0
            }
            st.session_state.goals = {
                'leads_goal': 100,
                'appointments_goal': 50,
                'closings_goal': 25
            }
            st.session_state.app_state.initialized = True
    except Exception as e:
        st.error(f"Error initializing session state: {str(e)}")

def save_state() -> None:
    try:
        if isinstance(st.session_state.data_state.historical_data, pd.DataFrame):
            st.session_state.data_state.historical_data = st.session_state.data_state.historical_data.to_dict('records')
        st.session_state.app_state = st.session_state.app_state.to_dict()
    except Exception as e:
        st.error(f"Error saving state: {str(e)}")

def restore_state() -> None:
    try:
        if isinstance(st.session_state.data_state.historical_data, list):
            st.session_state.data_state.historical_data = pd.DataFrame(st.session_state.data_state.historical_data)
        st.session_state.app_state = AppState.from_dict(st.session_state.app_state)
    except Exception as e:
        st.error(f"Error restoring state: {str(e)}")

@st.cache_data(ttl=3600)
def process_uploaded_file(uploaded_file) -> Optional[pd.DataFrame]:
    try:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)
        return df
    except Exception as e:
        st.error(f"Error processing file: {str(e)}")
        return None

# Tab rendering functions
def render_tab_input() -> None:
    st.markdown('<h2 class="section-header">Data Input</h2>', unsafe_allow_html=True)

    # Manual input section
    st.markdown("### Manual Data Entry")
    st.markdown("Enter your sales pipeline data below. Required fields are marked with an asterisk (*)")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown('<div class="required-field">', unsafe_allow_html=True)
        leads = st.number_input("Number of Leads", min_value=0, value=st.session_state.input_values.get('leads', 100))
        st.session_state.input_values['leads'] = leads

        appointments = st.number_input("Number of Appointments", min_value=0, max_value=leads,
                                         value=min(st.session_state.input_values.get('appointments', 50), leads))
        st.session_state.input_values['appointments'] = appointments
        st.markdown('</div>', unsafe_allow_html=True)

    with col2:
        st.markdown('<div class="required-field">', unsafe_allow_html=True)
        closings = st.number_input("Number of Closings", min_value=0, max_value=appointments,
                                    value=min(st.session_state.input_values.get('closings', 25), appointments))
        st.session_state.input_values['closings'] = closings

        st.markdown('<div class="optional-field">', unsafe_allow_html=True)
        cost = st.number_input("Total Cost ($)", min_value=0.0,
                               value=st.session_state.input_values.get('cost', 0.0))
        st.session_state.input_values['cost'] = cost
        st.markdown('</div>', unsafe_allow_html=True)

    # Sample data downloads
    st.markdown('<h3 class="section-header">Data Import</h3>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        sample_df = pd.DataFrame({
            'month': pd.date_range(start='2022-01-01', periods=12, freq='M'),
            'leads': np.random.randint(50, 150, 12),
            'appointments': np.random.randint(20, 100, 12),
            'closings': np.random.randint(5, 50, 12),
            'cost': np.random.uniform(2000, 10000, 12)
        })
        csv = sample_df.to_csv(index=False).encode('utf-8')
        st.download_button(label="📥 Download Sample CSV", data=csv, file_name='sample_data.csv', mime='text/csv')

    with col2:
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            sample_df.to_excel(writer, index=False)
        st.download_button(label="📥 Download Sample Excel", data=excel_buffer.getvalue(),
                           file_name='sample_data.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    uploaded_file = st.file_uploader("Upload Your Data File", type=["csv", "xlsx"],
                                       help="Upload your data file in CSV or Excel format")

    if uploaded_file is not None:
        with st.spinner("Processing file..."):
            df = process_uploaded_file(uploaded_file)
            if df is not None:
                st.session_state.data_state.historical_data = df
                show_success("File processed successfully! Now you can analyze the data.")

def render_tab_analysis() -> None:
    st.markdown('<h2 class="section-header">Analysis Dashboard</h2>', unsafe_allow_html=True)

    if not st.session_state.data_state.has_data():
        show_warning("No data available for analysis. Please add data in the input tab.")
        return

    df = st.session_state.data_state.historical_data

    # Date range filter
    st.sidebar.header("Analysis Filters")
    date_range = st.sidebar.date_input("Select Date Range", value=(df['month'].min(), df['month'].max()),
                                        min_value=df['month'].min(), max_value=df['month'].max())

    # Filter data
    mask = (df['month'].dt.date >= date_range[0]) & (df['month'].dt.date <= date_range[1])
    filtered_df = df.loc[mask]

    if not filtered_df.empty:
        # Show metrics
        metrics = calculate_metrics(filtered_df)
        create_metrics_dashboard(metrics)

        # Seasonality analysis
        st.header("Seasonality Analysis")
        metric_choice = st.selectbox("Select Metric for Seasonality Analysis:", ['leads', 'appointments', 'closings', 'cost'])

        if len(filtered_df) >= 24:
            with st.spinner("Calculating seasonality..."):
                fig, decomp = analyze_seasonality(filtered_df.set_index('month'), metric_choice)
                if fig is not None:
                    st.plotly_chart(fig, use_container_width=True)

                    # Show stationarity test results
                    stationarity = check_stationarity(filtered_df[metric_choice])
                    st.write("Stationarity Test Results:")
                    st.write(f"Test Statistic: {stationarity['test_statistic']:.2f}")
                    st.write(f"P-value: {stationarity['p_value']:.4f}")
                    st.write(f"Series is {'stationary' if stationarity['is_stationary'] else 'non-stationary'}")
        else:
            show_warning("Need at least 24 months of data for seasonality analysis")

def render_tab_forecasting() -> None:
    st.markdown('<h2 class="section-header">Forecast Analysis</h2>', unsafe_allow_html=True)

    if not st.session_state.data_state.has_data():
        show_warning("No data available for forecasting. Please add data in the input tab.")
        return

    # Model configuration
    st.subheader("Model Configuration")
    col1, col2 = st.columns(2)

    with col1:
        model_type = st.selectbox("Select Model Type", ['linear regression', 'ridge regression', 'lasso regression', 'random forest'])
        confidence_level = st.slider("Confidence Level", min_value=0.80, max_value=0.99, value=0.95, step=0.01)

    with col2:
        st.markdown("### Model Settings")
        st.markdown("""
        - **Linear Regression**: Simple and interpretable
        - **Ridge Regression**: Good for correlated features
        - **Lasso Regression**: Feature selection capabilities
        - **Random Forest**: Complex patterns, handles nonlinearity
        """)

    # Initialize forecaster and make predictions
    forecaster = EnhancedForecaster()

    with st.spinner("Generating forecast..."):
        df = st.session_state.data_state.historical_data
        try:
            forecaster.train(df.set_index('month'), model_type=model_type)

            # Cross-validation results
            cv_scores = forecaster.cross_validate(n_splits=5)
            if len(cv_scores) > 0:
                st.subheader("Cross-Validation Results")
                cv_df = pd.DataFrame({'split': range(1, len(cv_scores) + 1), 'r² score': cv_scores})
                st.bar_chart(cv_df.set_index('split'))
                st.write(f"Average r² score: {cv_scores.mean():.3f}")

            # Make predictions
            future_x = forecaster.prepare_features(df.set_index('month')).tail(1)
            predictions, lower_bound, upper_bound = forecaster.predict_with_intervals(future_x, confidence=confidence_level)

            if len(predictions) > 0:
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Forecasted Closings", f"{predictions[0]:.1f}")
                with col2:
                    st.metric("Lower Bound", f"{lower_bound[0]:.1f}")
                with col3:
                    st.metric("Upper Bound", f"{upper_bound[0]:.1f}")

                # Store forecasts in state
                st.session_state.data_state.forecasts = {
                    'predictions': predictions.tolist(),
                    'lower_bound': lower_bound.tolist(),
                    'upper_bound': upper_bound.tolist(),
                    'model_type': model_type,
                    'confidence_level': confidence_level
                }

                # Feature importance for random forest
                if model_type == 'random forest' and forecaster.feature_importance is not None:
                    st.subheader("Feature Importance")
                    fig = px.bar(forecaster.feature_importance, x='feature', y='importance', title='Feature Importance Analysis')
                    st.plotly_chart(fig, use_container_width=True)

        except Exception as e:
            show_error(f"Error in forecasting: {str(e)}")

def render_tab_export() -> None:
    st.markdown('<h2 class="section-header">Export Data</h2>', unsafe_allow_html=True)

    if not st.session_state.data_state.has_data():
        show_warning("No data available for export. Please add data in the input tab.")
        return

    # Export settings
    st.subheader("Export Settings")
    st.write("Select sections to include in the export:")

    col1, col2 = st.columns(2)
    with col1:
        include_sections = {
            'historical_data': st.checkbox("Historical Data", value=True),
            'metrics': st.checkbox("Metrics Analysis", value=True)
        }
    with col2:
        include_sections.update({
            'forecasts': st.checkbox("Forecasts", value=True),
            'metadata': st.checkbox("Metadata", value=True)
        })

    # Initialize exporter
    exporter = EnhancedExporter(st.session_state.data_state, st.session_state.app_state)

    # Export buttons
    st.subheader("Download Options")
    col1, col2 = st.columns(2)

    with col1:
        if st.button("Download Excel Report"):
            with st.spinner("Preparing Excel report..."):
                excel_data = exporter.create_excel_report([k for k, v in include_sections.items() if v])
                if excel_data:
                    st.download_button(label="📥 Download Excel", data=excel_data,
                                       file_name=f'sales_forecast_{exporter.timestamp}.xlsx',
                                       mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    show_success("Excel report generated successfully!")

    with col2:
        if st.button("Download Complete Package"):
            with st.spinner("Preparing export package..."):
                zip_data = exporter.create_zip_export([k for k, v in include_sections.items() if v])
                if zip_data:
                    st.download_button(label="📥 Download Zip", data=zip_data,
                                       file_name=f'sales_forecast_package_{exporter.timestamp}.zip',
                                       mime='application/zip')
                    show_success("Export package generated successfully!")

def main():
    """Main application function"""
    initialize_session_state()

    st.title("Sales Pipeline Analytics Dashboard")
    st.markdown("Advanced analytics and forecasting tool for sales pipeline management")

    tab1, tab2, tab3, tab4 = st.tabs([
        "📝 Input & Upload",
        "📊 Analysis",
        "📈 Forecasting",
        "💾 Export"
    ])

    with tab1:
        render_tab_input()
    with tab2:
        render_tab_analysis()
    with tab3:
        render_tab_forecasting()
    with tab4:
        render_tab_export()

if __name__ == "__main__":
    main()
